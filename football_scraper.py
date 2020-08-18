from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import time
import openpyxl
import re

PATH = "./chromedriver"

EXCEL_PATH = "./transfer_data.xlsx"

driver = webdriver.Chrome(PATH)

workbook = openpyxl.load_workbook(EXCEL_PATH)
sheet = workbook.active

player_num = []

# get team transfer data 2004-2020. (4, 20)
for i in range(4, 20):

    # setting the number of year.
    if i < 10:
        year = "0" + str(i)
    else:
        year = i

    # open the new window and move to it.
    driver.execute_script("window.open()")
    handle_array = driver.window_handles
    new_window = handle_array[1]
    driver.switch_to.window(new_window)

    # access to URL.
    driver.get(f'https://www.transfermarkt.com/tottenham-hotspur/transfers/verein/148/plus/1?saison_id=20{year}&pos=&detailpos=&w_s=')

    # correct data from page.
    try:
        season = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH , '//*[@id="main"]/div[11]/div/div[1]/h2'))
        ).text.split(' ')[1]

        print(season)

        trs = driver.find_elements_by_css_selector('#main > div:nth-child(18) > div > div:nth-child(3) > div.responsive-table > table > tbody > tr')

        row = len(player_num) + 2

        for tr in trs:

            personal_data = tr.text.splitlines()

            if len(personal_data) > 5:
                fee = re.findall(r"[-+]?\d*\.\d+|\d+", personal_data[5])
            
            else:
                fee = [];

            if len(fee):
                sheet.cell(row=row, column=1).value = season
                sheet.cell(row=row, column=2).value = personal_data[0]
                sheet.cell(row=row, column=3).value = personal_data[1]
                sheet.cell(row=row, column=4).value = personal_data[2].split(' ')[0]
                sheet.cell(row=row, column=5).value = personal_data[3]
                sheet.cell(row=row, column=6).value = personal_data[4]
                print(personal_data[5])
                if "m" in personal_data[5]:
                    mil = float(fee[0]) * 1000000
                    sheet.cell(row=row, column=7).value = mil

                elif "Th" in personal_data[5]:
                    th = float(fee[0]) * 1000
                    sheet.cell(row=row, column=7).value = th
                else:
                    sheet.cell(row=row, column=7).value = personal_data[5]

                player_num.append("0")
                row += 1

        time.sleep(3)

        # close current window, go back to the main window.
        driver.close()
        driver.switch_to.window(handle_array[0])

    except TimeoutException:
        print('取得失敗！')

    time.sleep(3)

workbook.save(EXCEL_PATH)

driver.quit()